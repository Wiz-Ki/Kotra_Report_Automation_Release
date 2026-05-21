from __future__ import annotations

from pathlib import Path

from compat import ensure_stdlib_selectors

ensure_stdlib_selectors()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from config import DEFAULT_INPUT_EXCEL
from field_mapping import ORIGINAL_EXAMPLE_ROWS, ORIGINAL_TEMPLATE_COLUMNS


def create_input_template(path: str | Path = DEFAULT_INPUT_EXCEL) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(ORIGINAL_EXAMPLE_ROWS, columns=ORIGINAL_TEMPLATE_COLUMNS)
    df.to_excel(path, index=False)
    apply_template_style(path)
    return path


def apply_template_style(path: Path) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    border_side = Side(style="thin", color="D9E2EC")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    column_widths = {
        "A": 12,
        "B": 18,
        "C": 18,
        "D": 34,
        "E": 24,
        "F": 22,
        "G": 34,
        "H": 28,
    }
    for column_letter, width in column_widths.items():
        sheet.column_dimensions[column_letter].width = width

    sheet.row_dimensions[1].height = 42

    for cell in sheet[1]:
        header = str(cell.value or "")
        font_color = "C00000" if "필수" in header else "666666"
        cell.font = Font(color=font_color, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            header = str(sheet.cell(1, cell.column).value or "")
            cell.font = Font(color="000000" if "필수" in header else "666666")
            cell.alignment = Alignment(horizontal="center" if cell.column in {1, 4, 5, 6, 8} else "left", vertical="center", wrap_text=True)
            cell.border = border

    workbook.save(path)
