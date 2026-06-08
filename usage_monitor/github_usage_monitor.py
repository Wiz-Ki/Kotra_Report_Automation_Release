#!/usr/bin/env python3
"""Private GitHub usage monitor.

This tool keeps GitHub traffic history in a separate local folder and exports
the result to an Excel workbook plus a small local HTML dashboard.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from html import escape
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - handled at runtime for user-friendly error
    Workbook = None


BASE_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = BASE_DIR.parent
DATA_PATH = BASE_DIR / "data" / "github_traffic_history.json"
EXCEL_PATH = BASE_DIR / "output" / "github_usage_report.xlsx"
DASHBOARD_PATH = BASE_DIR / "output" / "dashboard.html"
DEFAULT_REPO = "Wiz-Ki/Kotra_Report_Automation_Release"
GITHUB_API = "https://api.github.com"


class GitHubApiError(RuntimeError):
    pass


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def repo_from_remote(remote_name: str) -> str | None:
    try:
        result = subprocess.run(
            ["git", "config", "--get", f"remote.{remote_name}.url"],
            cwd=WORKSPACE_DIR,
            text=True,
            capture_output=True,
            check=False,
        )
    except OSError:
        return None

    url = result.stdout.strip()
    if not url:
        return None

    match = re.search(r"github\.com[:/](?P<owner>[^/]+)/(?P<repo>[^/.]+)(?:\.git)?$", url)
    if not match:
        return None
    return f"{match.group('owner')}/{match.group('repo')}"


def infer_repo() -> str:
    return (
        os.environ.get("GITHUB_TRAFFIC_REPOSITORY")
        or repo_from_remote("release")
        or repo_from_remote("origin")
        or DEFAULT_REPO
    )


def token_from_env() -> str | None:
    return os.environ.get("GITHUB_TOKEN") or os.environ.get("GH_TOKEN")


def google_sheets_webhook_from_env() -> str | None:
    return os.environ.get("GOOGLE_SHEETS_WEBHOOK_URL")


def google_sheets_secret_from_env() -> str | None:
    return os.environ.get("GOOGLE_SHEETS_WEBHOOK_SECRET")


def validate_repo(repo: str) -> str:
    repo = repo.strip()
    if not re.match(r"^[A-Za-z0-9_.-]+/[A-Za-z0-9_.-]+$", repo):
        raise ValueError(f"Invalid repository name: {repo!r}. Use OWNER/REPO.")
    return repo


def api_path(repo: str, suffix: str) -> str:
    owner, name = validate_repo(repo).split("/", 1)
    return f"/repos/{quote(owner)}/{quote(name)}{suffix}"


def github_get(path: str, token: str) -> Any:
    request = Request(
        f"{GITHUB_API}{path}",
        headers={
            "Accept": "application/vnd.github+json",
            "Authorization": f"Bearer {token}",
            "User-Agent": "kotra-private-github-usage-monitor",
            "X-GitHub-Api-Version": "2022-11-28",
        },
    )
    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise GitHubApiError(f"GitHub API returned {exc.code} for {path}: {body}") from exc
    except URLError as exc:
        raise GitHubApiError(f"Could not reach GitHub API for {path}: {exc}") from exc


def json_post(url: str, payload: dict[str, Any], headers: dict[str, str] | None = None) -> Any:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = Request(
        url,
        data=body,
        headers={
            "Content-Type": "application/json; charset=utf-8",
            "User-Agent": "kotra-private-github-usage-monitor",
            **(headers or {}),
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=30) as response:
            text = response.read().decode("utf-8")
            if not text:
                return {}
            try:
                result = json.loads(text)
            except json.JSONDecodeError as exc:
                preview = text[:1000].replace("\n", " ")
                raise RuntimeError(f"Cloud upload returned a non-JSON response: {preview}") from exc
            if isinstance(result, dict) and result.get("ok") is False:
                raise RuntimeError(f"Cloud upload failed: {result}")
            return result
    except HTTPError as exc:
        body_text = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Cloud upload returned {exc.code}: {body_text}") from exc
    except URLError as exc:
        raise RuntimeError(f"Could not reach cloud upload endpoint: {exc}") from exc


def load_history(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {
            "schema_version": 1,
            "repo": None,
            "release_repo": None,
            "created_at": utc_now(),
            "updated_at": None,
            "daily": {},
            "snapshots": [],
            "referrers": [],
            "paths": [],
            "releases": {
                "repo": None,
                "collected_at": None,
                "total_asset_downloads": 0,
                "assets": [],
            },
        }
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def save_history(path: Path, history: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(history, file, ensure_ascii=False, indent=2, sort_keys=True)
        file.write("\n")


def merge_daily(history: dict[str, Any], key_count: str, key_unique: str, items: list[dict[str, Any]]) -> None:
    daily = history.setdefault("daily", {})
    for item in items:
        day = item["timestamp"][:10]
        row = daily.setdefault(
            day,
            {
                "clones": 0,
                "unique_cloners": 0,
                "views": 0,
                "unique_visitors": 0,
            },
        )
        row[key_count] = int(item.get("count", 0))
        row[key_unique] = int(item.get("uniques", 0))


def fetch_release_downloads(repo: str, token: str) -> dict[str, Any]:
    assets: list[dict[str, Any]] = []
    page = 1
    while True:
        releases = github_get(api_path(repo, f"/releases?per_page=100&page={page}"), token)
        if not releases:
            break
        for release in releases:
            for asset in release.get("assets", []):
                assets.append(
                    {
                        "release": release.get("tag_name") or release.get("name") or "",
                        "asset": asset.get("name") or "",
                        "downloads": int(asset.get("download_count", 0)),
                        "published_at": release.get("published_at") or "",
                    }
                )
        if len(releases) < 100:
            break
        page += 1

    assets.sort(key=lambda item: item["downloads"], reverse=True)
    return {
        "repo": repo,
        "collected_at": utc_now(),
        "total_asset_downloads": sum(item["downloads"] for item in assets),
        "assets": assets,
    }


def fetch_and_merge(history: dict[str, Any], repo: str, release_repo: str | None, token: str) -> dict[str, Any]:
    now = utc_now()
    clones = github_get(api_path(repo, "/traffic/clones?per=day"), token)
    views = github_get(api_path(repo, "/traffic/views?per=day"), token)
    referrers = github_get(api_path(repo, "/traffic/popular/referrers"), token)
    paths = github_get(api_path(repo, "/traffic/popular/paths"), token)

    history["repo"] = repo
    history["release_repo"] = release_repo
    history["updated_at"] = now
    merge_daily(history, "clones", "unique_cloners", clones.get("clones", []))
    merge_daily(history, "views", "unique_visitors", views.get("views", []))
    history["referrers"] = referrers
    history["paths"] = paths
    history["snapshots"] = (history.get("snapshots", []) + [
        {
            "collected_at": now,
            "clones_14d": int(clones.get("count", 0)),
            "unique_cloners_14d": int(clones.get("uniques", 0)),
            "views_14d": int(views.get("count", 0)),
            "unique_visitors_14d": int(views.get("uniques", 0)),
        }
    ])[-730:]

    if release_repo:
        history["releases"] = fetch_release_downloads(release_repo, token)

    return history


def daily_rows(history: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for day, values in sorted(history.get("daily", {}).items()):
        rows.append(
            {
                "date": day,
                "clones": int(values.get("clones", 0)),
                "unique_cloners": int(values.get("unique_cloners", 0)),
                "views": int(values.get("views", 0)),
                "unique_visitors": int(values.get("unique_visitors", 0)),
            }
        )
    return rows


def history_summary(history: dict[str, Any]) -> dict[str, Any]:
    rows = daily_rows(history)
    recent = rows[-14:]
    releases = history.get("releases", {})
    return {
        "repo": history.get("repo") or infer_repo(),
        "updated_at": history.get("updated_at"),
        "tracked_days": len(rows),
        "clones_14d": sum(row["clones"] for row in recent),
        "unique_cloners_14d": sum(row["unique_cloners"] for row in recent),
        "views_14d": sum(row["views"] for row in recent),
        "unique_visitors_14d": sum(row["unique_visitors"] for row in recent),
        "release_asset_downloads": int(releases.get("total_asset_downloads", 0)),
    }


def cloud_payload(history: dict[str, Any], secret: str | None = None) -> dict[str, Any]:
    releases = history.get("releases", {})
    payload = {
        "schema_version": 1,
        "secret": secret,
        "summary": history_summary(history),
        "daily": daily_rows(history),
        "snapshots": history.get("snapshots", []),
        "referrers": history.get("referrers", []),
        "paths": history.get("paths", []),
        "releases": {
            "repo": releases.get("repo"),
            "collected_at": releases.get("collected_at"),
            "total_asset_downloads": int(releases.get("total_asset_downloads", 0)),
            "assets": releases.get("assets", []),
        },
    }
    if not secret:
        payload.pop("secret")
    return payload


def upload_to_google_sheets(history: dict[str, Any], webhook_url: str, secret: str | None = None) -> Any:
    return json_post(webhook_url, cloud_payload(history, secret))


def write_excel(history: dict[str, Any], path: Path) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required. Install it with: pip install openpyxl")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    white = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    rows = daily_rows(history)
    summary_values = history_summary(history)
    releases = history.get("releases", {})

    summary["A1"] = "GitHub Usage Monitor"
    summary["A1"].font = Font(size=18, bold=True, color="1F4E78")
    summary["A3"] = "Repository"
    summary["B3"] = history.get("repo") or "-"
    summary["A4"] = "Last Updated"
    summary["B4"] = history.get("updated_at") or "-"
    summary["A5"] = "Tracked Days"
    summary["B5"] = summary_values["tracked_days"]

    metrics = [
        ("14D Clones", summary_values["clones_14d"]),
        ("14D Unique Cloners", summary_values["unique_cloners_14d"]),
        ("14D Views", summary_values["views_14d"]),
        ("14D Unique Visitors", summary_values["unique_visitors_14d"]),
        ("Release Asset Downloads", summary_values["release_asset_downloads"]),
    ]
    for index, (label, value) in enumerate(metrics, start=7):
        summary[f"A{index}"] = label
        summary[f"B{index}"] = value
        summary[f"A{index}"].font = bold

    summary["A14"] = "Note"
    summary["B14"] = "GitHub traffic data is available for only the last 14 days, so this workbook preserves history from each collection run."
    summary["B15"] = "Unique values are GitHub's period-based unique counts and should be treated as directional indicators."

    daily = wb.create_sheet("Daily History")
    append_table(
        daily,
        ["Date", "Clones", "Unique Cloners", "Views", "Unique Visitors"],
        [[row["date"], row["clones"], row["unique_cloners"], row["views"], row["unique_visitors"]] for row in rows],
        header_fill,
    )

    if len(rows) >= 2:
        chart = LineChart()
        chart.title = "Daily GitHub Traffic"
        chart.y_axis.title = "Count"
        chart.x_axis.title = "Date"
        data = Reference(daily, min_col=2, max_col=5, min_row=1, max_row=len(rows) + 1)
        categories = Reference(daily, min_col=1, min_row=2, max_row=len(rows) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 9
        chart.width = 20
        daily.add_chart(chart, "G2")

    snapshots = wb.create_sheet("14D Snapshots")
    append_table(
        snapshots,
        ["Collected At", "Clones 14D", "Unique Cloners 14D", "Views 14D", "Unique Visitors 14D"],
        [
            [
                row.get("collected_at"),
                row.get("clones_14d", 0),
                row.get("unique_cloners_14d", 0),
                row.get("views_14d", 0),
                row.get("unique_visitors_14d", 0),
            ]
            for row in history.get("snapshots", [])
        ],
        header_fill,
    )

    referrers = wb.create_sheet("Referrers")
    append_table(
        referrers,
        ["Referrer", "Count", "Uniques"],
        [[row.get("referrer"), row.get("count", 0), row.get("uniques", 0)] for row in history.get("referrers", [])],
        header_fill,
    )

    paths = wb.create_sheet("Popular Paths")
    append_table(
        paths,
        ["Path", "Title", "Count", "Uniques"],
        [[row.get("path"), row.get("title"), row.get("count", 0), row.get("uniques", 0)] for row in history.get("paths", [])],
        header_fill,
    )

    assets = wb.create_sheet("Release Downloads")
    append_table(
        assets,
        ["Release", "Asset", "Downloads", "Published At"],
        [
            [row.get("release"), row.get("asset"), row.get("downloads", 0), row.get("published_at")]
            for row in releases.get("assets", [])
        ],
        header_fill,
    )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        autosize(sheet)
        for cell in sheet[1]:
            cell.fill = title_fill if sheet.title == "Summary" else header_fill
            cell.font = white if sheet.title == "Summary" else bold
            cell.alignment = Alignment(horizontal="center")

    wb.save(path)


def append_table(sheet: Any, headers: list[str], rows: list[list[Any]], header_fill: Any) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)


def autosize(sheet: Any) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 55)


def render_dashboard(history: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "repo": history.get("repo") or infer_repo(),
        "updated_at": history.get("updated_at"),
        "daily": daily_rows(history),
        "snapshots": history.get("snapshots", []),
        "referrers": history.get("referrers", []),
        "paths": history.get("paths", []),
        "releases": history.get("releases", {}),
    }
    embedded = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    repo_title = escape(str(data["repo"]))
    html = f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Private GitHub Usage Monitor - {repo_title}</title>
<style>
body{{margin:0;background:#f6f7f9;color:#1d2430;font-family:"Segoe UI","Malgun Gothic",Arial,sans-serif}}
main{{width:min(1120px,calc(100% - 32px));margin:0 auto;padding:28px 0 42px}}
header{{display:flex;justify-content:space-between;gap:18px;align-items:end;margin-bottom:18px}}
h1{{margin:0;font-size:28px;letter-spacing:0}} h2{{margin:0 0 12px;font-size:17px}}
.muted{{color:#657184;font-size:13px;line-height:1.5}} .repo{{color:#2563eb;font-weight:700}}
.grid{{display:grid;gap:14px}} .cards{{grid-template-columns:repeat(5,minmax(0,1fr));margin-bottom:14px}}
.panels{{grid-template-columns:minmax(0,1.4fr) minmax(320px,.8fr)}} .lower{{grid-template-columns:repeat(3,minmax(0,1fr));margin-top:14px}}
.card,.panel{{background:white;border:1px solid #d9dee7;border-radius:8px;box-shadow:0 1px 2px rgba(15,23,42,.04)}}
.card{{min-height:108px;padding:15px;display:flex;flex-direction:column;justify-content:space-between}}
.label{{color:#657184;font-size:12px;font-weight:700;text-transform:uppercase}} .value{{font-size:30px;font-weight:800;color:#111827}}
.hint{{color:#657184;font-size:12px}} .panel{{padding:18px;min-width:0}} .chart{{height:320px;border:1px solid #d9dee7;border-radius:8px;background:#fbfcfe}}
svg{{width:100%;height:100%;display:block}} table{{width:100%;border-collapse:collapse;table-layout:fixed}}
th,td{{padding:9px 0;border-bottom:1px solid #edf0f5;text-align:left;font-size:13px;vertical-align:middle;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
th{{color:#657184;font-size:12px}} .num{{text-align:right;font-variant-numeric:tabular-nums}} .empty{{display:grid;place-items:center;min-height:180px;color:#657184;text-align:center;line-height:1.6}}
.notes{{margin-top:15px;color:#657184;font-size:12px;line-height:1.6}} code{{background:#eef2f7;border:1px solid #d8deea;padding:2px 5px;border-radius:5px}}
@media(max-width:940px){{header{{display:block}}.cards{{grid-template-columns:repeat(2,minmax(0,1fr))}}.panels,.lower{{grid-template-columns:1fr}}}}
@media(max-width:520px){{main{{width:min(100% - 20px,1120px)}}.cards{{grid-template-columns:1fr}}h1{{font-size:23px}}}}
</style>
</head>
<body>
<main>
<header><div><h1>Private GitHub Usage Monitor</h1><div class="muted">대상 저장소 <span class="repo" id="repo"></span></div></div><div class="muted" id="updated"></div></header>
<section class="grid cards" id="cards"></section>
<section class="grid panels"><article class="panel"><h2>일별 추세</h2><div class="chart" id="chart"></div></article><article class="panel"><h2>최근 수집 스냅샷</h2><div id="snapshots"></div></article></section>
<section class="grid lower"><article class="panel"><h2>유입 경로</h2><div id="referrers"></div></article><article class="panel"><h2>인기 페이지</h2><div id="paths"></div></article><article class="panel"><h2>릴리즈 다운로드</h2><div id="releases"></div></article></section>
<p class="notes">이 파일은 로컬 전용입니다. GitHub Traffic API는 최근 14일만 제공하므로, 매일 실행할수록 장기 추세가 정확해집니다. <code>Code &gt; Download ZIP</code> 다운로드 수는 GitHub가 별도 장기 지표로 제공하지 않습니다.</p>
</main>
<script id="data" type="application/json">{embedded}</script>
<script>
const data=JSON.parse(document.getElementById('data').textContent),fmt=new Intl.NumberFormat('ko-KR'),dt=new Intl.DateTimeFormat('ko-KR',{{dateStyle:'medium',timeStyle:'short'}});
const esc=s=>String(s??'').replace(/[&<>"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
const sum=(rows,k)=>rows.reduce((a,r)=>a+Number(r[k]||0),0),recent=data.daily.slice(-14);
document.getElementById('repo').textContent=data.repo||'-';
document.getElementById('updated').textContent=data.updated_at?'마지막 업데이트: '+dt.format(new Date(data.updated_at)):'아직 수집된 데이터가 없습니다';
function card(l,v,h){{return `<article class="card"><div class="label">${{l}}</div><div class="value">${{fmt.format(v)}}</div><div class="hint">${{h}}</div></article>`}}
document.getElementById('cards').innerHTML=[
card('14D Clones',sum(recent,'clones'),'최근 14일 clone 총량'),
card('14D Unique',sum(recent,'unique_cloners'),'최근 14일 일별 unique 합계'),
card('14D Views',sum(recent,'views'),'최근 14일 방문 총량'),
card('Release Downloads',Number(data.releases?.total_asset_downloads||0),'릴리즈 asset 누적'),
card('Tracked Days',data.daily.length,'저장된 일별 데이터')
].join('');
function empty(m){{return `<div class="empty">${{m}}</div>`}}
function chart(rows){{if(!rows.length)return empty('아직 표시할 데이터가 없습니다.');const w=900,h=320,p={{l:42,r:18,t:22,b:38}},series=[['clones','#2563eb'],['unique_cloners','#0f8b6f'],['views','#b7791f'],['unique_visitors','#c2415d']],max=Math.max(1,...rows.flatMap(r=>series.map(([k])=>Number(r[k]||0)))),x=i=>p.l+(rows.length===1?0:i*(w-p.l-p.r)/(rows.length-1)),y=v=>h-p.b-(Number(v||0)/max)*(h-p.t-p.b);let g=[0,.25,.5,.75,1].map(t=>{{const v=Math.round(max*t),yy=y(v);return `<line x1="${{p.l}}" y1="${{yy}}" x2="${{w-p.r}}" y2="${{yy}}" stroke="#e8edf5"/><text x="8" y="${{yy+4}}" fill="#657184" font-size="11">${{fmt.format(v)}}</text>`}}).join('');let lines=series.map(([k,c])=>`<path d="${{rows.map((r,i)=>(i?'L':'M')+' '+x(i).toFixed(1)+' '+y(r[k]).toFixed(1)).join(' ')}}" fill="none" stroke="${{c}}" stroke-width="2.4" stroke-linecap="round" stroke-linejoin="round"/>`).join('');let labels=rows.filter((_,i)=>i===0||i===rows.length-1||i%Math.ceil(rows.length/6)===0).map(r=>{{const i=rows.indexOf(r);return `<text x="${{x(i)}}" y="${{h-13}}" fill="#657184" font-size="11" text-anchor="${{i===0?'start':i===rows.length-1?'end':'middle'}}">${{r.date.slice(5)}}</text>`}}).join('');return `<svg viewBox="0 0 ${{w}} ${{h}}">${{g}}${{lines}}${{labels}}</svg>`}}
document.getElementById('chart').innerHTML=chart(data.daily);
function table(rows,cols){{if(!rows.length)return empty('아직 데이터가 없습니다.');return `<table><thead><tr>${{cols.map(c=>`<th class="${{c.num?'num':''}}">${{c.l}}</th>`).join('')}}</tr></thead><tbody>${{rows.map(r=>`<tr>${{cols.map(c=>`<td class="${{c.num?'num':''}}" title="${{esc(c.f?c.f(r):r[c.k])}}">${{esc(c.f?c.f(r):r[c.k])}}</td>`).join('')}}</tr>`).join('')}}</tbody></table>`}}
document.getElementById('snapshots').innerHTML=table([...(data.snapshots||[])].slice(-8).reverse(),[{{l:'수집 시각',f:r=>dt.format(new Date(r.collected_at))}},{{l:'Clones',k:'clones_14d',num:true}},{{l:'Unique',k:'unique_cloners_14d',num:true}},{{l:'Views',k:'views_14d',num:true}}]);
document.getElementById('referrers').innerHTML=table((data.referrers||[]).slice(0,8),[{{l:'Referrer',k:'referrer'}},{{l:'Total',k:'count',num:true}},{{l:'Unique',k:'uniques',num:true}}]);
document.getElementById('paths').innerHTML=table((data.paths||[]).slice(0,8),[{{l:'Path',k:'path'}},{{l:'Total',k:'count',num:true}},{{l:'Unique',k:'uniques',num:true}}]);
document.getElementById('releases').innerHTML=table((data.releases?.assets||[]).slice(0,8),[{{l:'Asset',k:'asset'}},{{l:'Release',k:'release'}},{{l:'Downloads',k:'downloads',num:true}}]);
</script>
</body>
</html>"""
    path.write_text(html, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Collect private GitHub usage stats into Excel and HTML.")
    parser.add_argument("--repo", default=infer_repo(), help="Repository to monitor, in OWNER/REPO format.")
    parser.add_argument(
        "--release-repo",
        default=os.environ.get("GITHUB_RELEASE_REPOSITORY") or infer_repo(),
        help="Repository to use for release asset download counts. Use '' to disable.",
    )
    parser.add_argument("--data", type=Path, default=DATA_PATH, help="JSON history path.")
    parser.add_argument("--excel", type=Path, default=EXCEL_PATH, help="Excel output path.")
    parser.add_argument("--dashboard", type=Path, default=DASHBOARD_PATH, help="HTML dashboard output path.")
    parser.add_argument(
        "--google-sheets-webhook",
        default=google_sheets_webhook_from_env(),
        help="Google Apps Script web app URL. Defaults to GOOGLE_SHEETS_WEBHOOK_URL.",
    )
    parser.add_argument(
        "--google-sheets-secret",
        default=google_sheets_secret_from_env(),
        help="Optional shared secret. Defaults to GOOGLE_SHEETS_WEBHOOK_SECRET.",
    )
    parser.add_argument("--skip-cloud-upload", action="store_true", help="Do not upload to Google Sheets.")
    parser.add_argument("--no-fetch", action="store_true", help="Generate Excel/HTML from existing history only.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo = validate_repo(args.repo)
    release_repo = args.release_repo.strip() if isinstance(args.release_repo, str) else args.release_repo
    release_repo = validate_repo(release_repo) if release_repo else None

    history = load_history(args.data)
    if not args.no_fetch:
        token = token_from_env()
        if not token:
            print("GITHUB_TOKEN or GH_TOKEN is required. Set a GitHub token first.", file=sys.stderr)
            return 2
        history = fetch_and_merge(history, repo, release_repo, token)
        save_history(args.data, history)

    write_excel(history, args.excel)
    render_dashboard(history, args.dashboard)
    if args.google_sheets_webhook and not args.skip_cloud_upload:
        response = upload_to_google_sheets(history, args.google_sheets_webhook, args.google_sheets_secret)
        print(f"Google Sheets upload: {response}")
    elif not args.skip_cloud_upload:
        print("Google Sheets upload: skipped. Set GOOGLE_SHEETS_WEBHOOK_URL to enable cloud backup.")
    print(f"History:   {args.data}")
    print(f"Excel:     {args.excel}")
    print(f"Dashboard: {args.dashboard}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
